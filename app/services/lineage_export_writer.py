from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

SUMMARY_HEADERS = ["key", "value"]
TABLES_HEADERS = ["full_name", "short_name", "layer", "field_count", "system_name", "is_external"]
TABLE_LINEAGES_HEADERS = ["source_table", "target_table", "procedure", "data_source", "operation_type"]
FIELD_MAPPINGS_HEADERS = [
    "source_table",
    "source_column",
    "target_table",
    "target_column",
    "procedure",
    "transform_logic",
    "condition_metadata",
]


def build_lineage_export_workbook(export: dict[str, Any]) -> bytes:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    _write_rows(
        summary_ws,
        SUMMARY_HEADERS,
        [{"key": key, "value": value} for key, value in export.get("summary", {}).items()],
    )
    _write_rows(wb.create_sheet("Tables"), TABLES_HEADERS, export.get("tables", []))
    _write_rows(wb.create_sheet("TableLineages"), TABLE_LINEAGES_HEADERS, export.get("table_lineages", []))
    _write_rows(wb.create_sheet("FieldMappings"), FIELD_MAPPINGS_HEADERS, export.get("field_mappings", []))

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _write_rows(ws: Worksheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    _style_header(ws)
    for row in rows:
        ws.append([_cell_value(row.get(header)) for header in headers])
    _fit_columns(ws, headers, rows)
    ws.freeze_panes = "A2"


def _style_header(ws: Worksheet) -> None:
    fill = PatternFill("solid", fgColor="E6EEF8")
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _fit_columns(ws: Worksheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    for index, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in rows:
            value = _cell_value(row.get(header))
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = min(max_len + 2, 60)


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (bool, int, float, str)):
        return value
    return str(value)
