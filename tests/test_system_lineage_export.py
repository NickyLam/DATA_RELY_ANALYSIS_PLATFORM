from __future__ import annotations

from io import BytesIO
from unittest.mock import MagicMock

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.dependencies import get_lineage_service
from app.main import app
from app.services.lineage_service import LineageService


class _Parser:
    data_generation = 7

    def __init__(self, data: dict | None):
        self._data = data
        self.parse_existing_called = False

    def get_current_data(self) -> dict | None:
        return self._data

    def get_data_mtime(self) -> float:
        return 1710000000.0

    def parse_existing_data(self, *args, **kwargs):
        self.parse_existing_called = True
        raise AssertionError("export must not trigger parsing")


class _Cache:
    size = 0

    def build_index(self, *args, **kwargs):
        return None

    def clear(self):
        return None

    def get(self, *args, **kwargs):
        return None

    def set(self, *args, **kwargs):
        return None

    def search_by_keyword(self, *args, **kwargs):
        return []


def _sample_lineage_data() -> dict:
    return {
        "metadata": {"total_tables": 4, "total_table_lineages": 3, "total_field_mappings": 2},
        "tables": [
            {
                "full_name": "IDL.EDW_ORDER",
                "table_name": "EDW_ORDER",
                "data_source": "edw",
                "columns": [{"name": "ORDER_ID"}, {"name": "AMT"}],
            },
            {
                "full_name": "EDW_MART.EDW_FACT",
                "table_name": "EDW_FACT",
                "columns": [{"name": "ORDER_ID"}],
            },
            {
                "full_name": "RRP_ODS.RRP_SRC",
                "table_name": "RRP_SRC",
                "data_source": "rrp",
                "columns": [{"name": "ORDER_ID"}],
            },
            {
                "full_name": "MCS_IDL.MCS_RPT",
                "table_name": "MCS_RPT",
                "data_source": "mcs",
                "columns": [{"name": "ORDER_ID"}],
            },
        ],
        "procedures": [],
        "table_lineages": [
            {
                "source_table": "RRP_ODS.RRP_SRC",
                "target_table": "IDL.EDW_ORDER",
                "procedure": "RRP.P_LOAD_ORDER",
                "data_source": "edw",
                "operation_type": "INSERT",
            },
            {
                "source_table": "IDL.EDW_ORDER",
                "target_table": "MCS_IDL.MCS_RPT",
                "procedure": "EDW.P_PUSH_RPT",
                "data_source": "edw",
                "operation_type": "MERGE",
            },
            {
                "source_table": "MCS_IDL.MCS_RPT",
                "target_table": "RRP_ODS.RRP_SRC",
                "procedure": "MCS.P_IGNORE",
                "data_source": "mcs",
            },
        ],
        "field_mappings": [
            {
                "source_table": "RRP_ODS.RRP_SRC",
                "source_column": "ORDER_ID",
                "target_table": "IDL.EDW_ORDER",
                "target_column": "ORDER_ID",
                "procedure": "RRP.P_LOAD_ORDER",
                "transform_logic": "CAST(ORDER_ID AS VARCHAR2(32))",
                "where_conditions": [{"raw_text": "ORDER_ID IS NOT NULL"}],
            },
            {
                "source_table": "IDL.EDW_ORDER",
                "source_column": "ORDER_ID",
                "target_table": "MCS_IDL.MCS_RPT",
                "target_column": "ORDER_ID",
                "procedure": "EDW.P_PUSH_RPT",
                "join_conditions": [{"raw_text": "A.ORDER_ID = B.ORDER_ID"}],
            },
        ],
        "caliber_infos": [],
    }


def test_lineage_service_exports_system_rows_and_external_nodes():
    parser = _Parser(_sample_lineage_data())
    service = LineageService(parser, _Cache())

    export = service.export_system_full_lineage("edw")

    table_names = {row["full_name"]: row for row in export["tables"]}
    assert table_names["IDL.EDW_ORDER"]["is_external"] is False
    assert table_names["EDW_MART.EDW_FACT"]["is_external"] is False
    assert table_names["RRP_ODS.RRP_SRC"]["is_external"] is True
    assert table_names["MCS_IDL.MCS_RPT"]["is_external"] is True
    assert {row["procedure"] for row in export["table_lineages"]} == {"RRP.P_LOAD_ORDER", "EDW.P_PUSH_RPT"}
    assert len(export["field_mappings"]) == 2
    assert export["summary"]["system_name"] == "edw"
    assert export["summary"]["total_nodes"] == 4
    assert export["summary"]["system_nodes"] == 2
    assert export["summary"]["external_nodes"] == 2
    assert export["summary"]["parser_generation"] == 7
    assert parser.parse_existing_called is False


def test_lineage_service_export_rejects_unknown_system_and_no_data():
    service = LineageService(_Parser(_sample_lineage_data()), _Cache())

    with pytest.raises(ValueError, match="unknown system"):
        service.export_system_full_lineage("unknown")

    empty_service = LineageService(_Parser(None), _Cache())
    with pytest.raises(ValueError, match="no parsed data"):
        empty_service.export_system_full_lineage("edw")


def test_xlsx_writer_creates_required_workbook_sheets_and_headers():
    from app.services.lineage_export_writer import (
        FIELD_MAPPINGS_HEADERS,
        SUMMARY_HEADERS,
        TABLE_LINEAGES_HEADERS,
        TABLES_HEADERS,
        build_lineage_export_workbook,
    )

    service = LineageService(_Parser(_sample_lineage_data()), _Cache())
    workbook_bytes = build_lineage_export_workbook(service.export_system_full_lineage("edw"))

    wb = openpyxl.load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    assert wb.sheetnames == ["Summary", "Tables", "TableLineages", "FieldMappings"]
    assert [cell.value for cell in next(wb["Summary"].iter_rows(max_row=1))] == SUMMARY_HEADERS
    assert [cell.value for cell in next(wb["Tables"].iter_rows(max_row=1))] == TABLES_HEADERS
    assert [cell.value for cell in next(wb["TableLineages"].iter_rows(max_row=1))] == TABLE_LINEAGES_HEADERS
    assert [cell.value for cell in next(wb["FieldMappings"].iter_rows(max_row=1))] == FIELD_MAPPINGS_HEADERS
    assert wb["Tables"].max_row == 5
    assert wb["TableLineages"].max_row == 3
    assert wb["FieldMappings"].max_row == 3


def test_system_lineage_export_api_downloads_xlsx_and_documents_openapi():
    service = MagicMock()
    service.export_system_full_lineage.return_value = {
        "summary": {
            "system_name": "edw",
            "generated_at": "2026-06-25T00:00:00",
            "data_mtime": 1710000000.0,
            "parser_generation": 7,
            "total_nodes": 1,
            "system_nodes": 1,
            "external_nodes": 0,
            "table_lineages": 0,
            "field_mappings": 0,
        },
        "tables": [
            {
                "full_name": "IDL.EDW_ORDER",
                "short_name": "EDW_ORDER",
                "layer": "idl",
                "field_count": 2,
                "system_name": "edw",
                "is_external": False,
            }
        ],
        "table_lineages": [],
        "field_mappings": [],
    }
    app.dependency_overrides[get_lineage_service] = lambda: service
    try:
        client = TestClient(app)
        response = client.get("/api/systems/edw/lineage/export")
        assert response.status_code == 200
        assert response.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in response.headers["content-disposition"]
        assert response.headers["content-disposition"].endswith(".xlsx\"")

        wb = openpyxl.load_workbook(BytesIO(response.content), read_only=True, data_only=True)
        assert wb.sheetnames == ["Summary", "Tables", "TableLineages", "FieldMappings"]

        schema = client.get("/openapi.json").json()
        content = schema["paths"]["/api/systems/{system_name}/lineage/export"]["get"]["responses"]["200"]["content"]
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in content
    finally:
        app.dependency_overrides.pop(get_lineage_service, None)


def test_system_lineage_export_api_maps_service_errors_to_404():
    service = MagicMock()
    service.export_system_full_lineage.side_effect = ValueError("unknown system: missing")
    app.dependency_overrides[get_lineage_service] = lambda: service
    try:
        response = TestClient(app).get("/api/systems/missing/lineage/export")
        assert response.status_code == 404
        assert "unknown system" in response.json()["detail"]
    finally:
        app.dependency_overrides.pop(get_lineage_service, None)
