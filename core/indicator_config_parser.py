"""Parse Excel config files and .proc stored procedure files for the financial indicator system."""

from __future__ import annotations

import logging
import re
import time
from pathlib import Path

import openpyxl

from core.indicator_models import (
    IndicatorCalcBase,
    IndicatorCalcGL,
    IndicatorConfigResult,
    IndicatorRel,
    ProcedureIndicatorInfo,
)
from core.indicator_sql_parser import IndicatorSQLParser

logger = logging.getLogger(__name__)

_EXCEL_FILENAME = "财务指标查询.xlsx"
_BASE_CALC_SHEET = "基础指标算法表"
_GL_CALC_SHEET = "总账指标算法表"

_BASE_CALC_COLUMNS = (
    "指标编号",
    "指标度量",
    "指标类型",
    "算法类型",
    "序列",
    "目标表名称",
    "源表名",
    "度量SQL",
    "CONDITION_SQL",
    "SQLCC",
    "START_DT",
    "END_DT",
    "UPDATE_USER",
    "INDEX_FLAG",
)

_GL_CALC_COLUMNS = (
    "指标号",
    "指标类型",
    "指标度量",
    "符号",
    "科目号",
    "LENGTH_VAL",
    "AMT_VAL",
    "START_DT",
    "END_DT",
)

_PROC_STEP_MAP: dict[str, int] = {
    "PRO_F_INDEX_CALC_GL": 1,
    "PRO_F_INDEX_CALC_BASE": 2,
    "PRO_F_INDEX_CALC_BRCHSUM": 3,
    "PRO_F_INDEX_CALC_DERIVE": 4,
}

_PROC_FILE_PATTERN = re.compile(r"PRO_F_INDEX_CALC_\w+\.proc$", re.IGNORECASE)
_PROC_HEADER_PATTERN = re.compile(
    r"CREATE\s+(?:OR\s+REPLACE\s+)?PROCEDURE\s+([\w.]+)",
    re.IGNORECASE,
)
_FROM_JOIN_PATTERN = re.compile(
    r"(?:FROM|JOIN)\s+([A-Za-z_][\w$#]*(?:\.[A-Za-z_][\w$#]*)?)",
    re.IGNORECASE,
)
_INSERT_INTO_PATTERN = re.compile(
    r"INSERT\s+INTO\s+([A-Za-z_][\w$#]*(?:\.[A-Za-z_][\w$#]*)?)",
    re.IGNORECASE,
)
_SKIP_TABLE_PREFIXES = (
    "FDL_IDX_LOG_",
    "SEQ_",
    "DUAL",
)
_CONFIG_TABLE_PATTERN = re.compile(
    r"FROM\s+(FDL_IDX_PARA_\w+)",
    re.IGNORECASE,
)


def _safe_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_file_with_encoding(file_path: Path) -> str:
    for encoding in ("utf-8", "gbk"):
        try:
            return file_path.read_text(encoding=encoding)
        except (UnicodeDecodeError, UnicodeError):
            continue
    return file_path.read_text(encoding="utf-8", errors="replace")


def _is_skip_table(table_name: str) -> bool:
    upper = table_name.upper()
    return any(upper.startswith(prefix) for prefix in _SKIP_TABLE_PREFIXES)


class IndicatorConfigParser:

    def __init__(self, data_path: Path | str) -> None:
        self.data_path = Path(data_path)  # ★ 修复：确保 data_path 始终为 Path 对象
        self._sql_parser = IndicatorSQLParser()

    def parse_all(self) -> IndicatorConfigResult:
        start = time.time()
        result = IndicatorConfigResult()

        try:
            result.base_calcs = self._parse_base_calc_excel()
        except Exception:
            logger.exception("解析基础指标算法表失败")
            result.base_calcs = []

        try:
            result.gl_calcs = self._parse_gl_calc_excel()
        except Exception:
            logger.exception("解析总账指标算法表失败")
            result.gl_calcs = []

        try:
            result.relations = self._parse_relations_from_base(result.base_calcs)
        except Exception:
            logger.exception("解析指标依赖关系失败")
            result.relations = []

        try:
            result.procedures = self._parse_proc_files()
        except Exception:
            logger.exception("解析存储过程文件失败")
            result.procedures = {}

        elapsed_sec = (time.time() - start)
        result.parse_time_sec = round(elapsed_sec, 3)
        logger.info(
            "指标配置解析完成: %d 基础指标, %d 总账指标, %d 依赖关系, %d 存储过程, 耗时 %.1f ms",
            len(result.base_calcs),
            len(result.gl_calcs),
            len(result.relations),
            len(result.procedures),
            elapsed_sec * 1000,
        )
        return result

    def _resolve_excel_path(self) -> Path:
        candidates = [
            self.data_path / _EXCEL_FILENAME,
            self.data_path / "config" / _EXCEL_FILENAME,
        ]
        for path in candidates:
            if path.exists():
                return path
        return candidates[0]

    def _parse_base_calc_excel(self) -> list[IndicatorCalcBase]:
        excel_path = self._resolve_excel_path()
        if not excel_path.exists():
            logger.warning("Excel 文件不存在: %s (已搜索: %s)", excel_path, ", ".join(str(p) for p in [self.data_path / _EXCEL_FILENAME, self.data_path / "config" / _EXCEL_FILENAME]))
            return []

        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
        try:
            if _BASE_CALC_SHEET not in wb.sheetnames:
                logger.warning("工作表不存在: %s", _BASE_CALC_SHEET)
                return []

            ws = wb[_BASE_CALC_SHEET]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return []

            header = [_safe_str(c) for c in rows[0]]
            col_index = self._build_col_index(header, _BASE_CALC_COLUMNS)

            results: list[IndicatorCalcBase] = []
            for row in rows[1:]:
                if not row or not row[0]:
                    continue

                calc = IndicatorCalcBase(
                    index_no=_safe_str(row[col_index.get("指标编号", 0)]),
                    index_measure=_safe_str(row[col_index.get("指标度量", 1)]),
                    algo_type=_safe_str(row[col_index.get("算法类型", 3)]),
                    call_level=_safe_str(row[col_index.get("序列", 4)]),
                    trg_table_name=_safe_str(row[col_index.get("目标表名称", 5)]),
                    src_table_name=_safe_str(row[col_index.get("源表名", 6)]),
                    measure_sql=_safe_str(row[col_index.get("度量SQL", 7)]),
                    condition_sql=_safe_str(row[col_index.get("CONDITION_SQL", 8)]),
                    sqlcc=_safe_str(row[col_index.get("SQLCC", 9)]),
                    index_flag=_safe_str(row[col_index.get("INDEX_FLAG", 13)]),
                    start_dt=_safe_str(row[col_index.get("START_DT", 10)]),
                    end_dt=_safe_str(row[col_index.get("END_DT", 11)]),
                )

                if calc.sqlcc:
                    extracted = self._sql_parser.extract_source_tables(calc.sqlcc)
                    if extracted and not calc.src_table_name:
                        calc.src_table_name = ",".join(extracted)
                    elif extracted:
                        existing = {
                            t.strip().upper()
                            for t in calc.src_table_name.split(",")
                            if t.strip()
                        }
                        for tbl in extracted:
                            if tbl.upper() not in existing:
                                calc.src_table_name = (
                                    calc.src_table_name.rstrip(",")
                                    + ","
                                    + tbl
                                )

                results.append(calc)

            logger.info("解析基础指标算法表: %d 条记录", len(results))
            return results
        finally:
            wb.close()

    def _parse_gl_calc_excel(self) -> list[IndicatorCalcGL]:
        excel_path = self._resolve_excel_path()
        if not excel_path.exists():
            logger.warning("Excel 文件不存在: %s", excel_path)
            return []

        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
        try:
            if _GL_CALC_SHEET not in wb.sheetnames:
                logger.warning("工作表不存在: %s", _GL_CALC_SHEET)
                return []

            ws = wb[_GL_CALC_SHEET]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return []

            header = [_safe_str(c) for c in rows[0]]
            col_index = self._build_col_index(header, _GL_CALC_COLUMNS)

            results: list[IndicatorCalcGL] = []
            for row in rows[1:]:
                if not row or not row[0]:
                    continue

                index_no_raw = _safe_str(row[col_index.get("指标号", 0)])
                if not index_no_raw or not re.match(r"^[A-Za-z0-9]+$", index_no_raw):
                    continue

                sign_raw = _safe_str(row[col_index.get("符号", 3)])
                sign_no = -1 if sign_raw.strip() == "-" else 1

                length_raw = _safe_str(row[col_index.get("LENGTH_VAL", 5)])
                try:
                    length_val = int(float(length_raw)) if length_raw else 0
                except (ValueError, TypeError):
                    length_val = 0

                calc = IndicatorCalcGL(
                    index_no=_safe_str(row[col_index.get("指标号", 0)]),
                    index_measure=_safe_str(row[col_index.get("指标度量", 2)]),
                    sign_no=sign_no,
                    subj_no=_safe_str(row[col_index.get("科目号", 4)]),
                    length_val=length_val,
                    amt_val=_safe_str(row[col_index.get("AMT_VAL", 6)]),
                    start_dt=_safe_str(row[col_index.get("START_DT", 7)]),
                    end_dt=_safe_str(row[col_index.get("END_DT", 8)]),
                )
                results.append(calc)

            logger.info("解析总账指标算法表: %d 条记录", len(results))
            return results
        finally:
            wb.close()

    def _parse_relations_from_base(
        self, base_calcs: list[IndicatorCalcBase]
    ) -> list[IndicatorRel]:
        deps_map: dict[str, set[str]] = {}

        for calc in base_calcs:
            depend_nos = self._extract_depend_index_nos(calc)
            if not depend_nos:
                continue

            if calc.index_no not in deps_map:
                deps_map[calc.index_no] = set()
            deps_map[calc.index_no].update(depend_nos)

        results: list[IndicatorRel] = []
        for index_no, depend_set in sorted(deps_map.items()):
            results.append(
                IndicatorRel(
                    index_no=index_no,
                    depend_index_nos=sorted(depend_set),
                )
            )

        logger.info("解析指标依赖关系: %d 组依赖", len(results))
        return results

    def _extract_depend_index_nos(self, calc: IndicatorCalcBase) -> list[str]:
        deps: set[str] = set()

        # 1. 从 condition_sql 提取（保留原有逻辑）
        cond = calc.condition_sql.strip()
        if cond:
            if calc.algo_type == "1":
                deps.update(self._split_comma_index_nos(cond))
            elif calc.algo_type == "3":
                deps.update(self._split_hash_index_nos(cond))
            elif calc.algo_type in ("5", "6"):
                deps.update(self._split_comma_index_nos(cond))

        # 2. 从 sqlcc 提取硬编码指标编号
        sqlcc = calc.sqlcc.strip()
        if sqlcc:
            pattern = re.compile(r"['\"]([A-Z]{2,3}\d{6,})['\"]", re.IGNORECASE)
            for match in pattern.finditer(sqlcc):
                ref_no = match.group(1).upper()
                if ref_no != calc.index_no.upper():
                    deps.add(ref_no)

        # 3. 从 measure_sql 提取硬编码指标编号
        measure = calc.measure_sql.strip()
        if measure:
            pattern = re.compile(r"['\"]([A-Z]{2,3}\d{6,})['\"]", re.IGNORECASE)
            for match in pattern.finditer(measure):
                ref_no = match.group(1).upper()
                if ref_no != calc.index_no.upper():
                    deps.add(ref_no)

        return sorted(deps)

    def _split_comma_index_nos(self, condition_sql: str) -> list[str]:
        parts = [p.strip() for p in condition_sql.split(",")]
        return [p for p in parts if p]

    def _split_hash_index_nos(self, condition_sql: str) -> list[str]:
        parts = [p.strip() for p in condition_sql.split("#")]
        return [p for p in parts if p]

    def _parse_proc_files(self) -> dict[str, ProcedureIndicatorInfo]:
        results: dict[str, ProcedureIndicatorInfo] = {}

        if not self.data_path.exists():
            logger.warning("数据目录不存在: %s", self.data_path)
            return results

        # ★ 优化：递归搜索所有 .proc 文件（不再仅限根目录）
        for proc_file in sorted(self.data_path.rglob("*.proc")):
            if not _PROC_FILE_PATTERN.search(proc_file.name):
                continue

            try:
                info = self._parse_single_proc(proc_file)
                if info.proc_name:
                    results[info.proc_name] = info
            except Exception:
                logger.exception("解析存储过程文件失败: %s", proc_file)
                continue

        logger.info("解析存储过程文件: %d 个", len(results))
        return results

    def _parse_single_proc(self, file_path: Path) -> ProcedureIndicatorInfo:
        content = _read_file_with_encoding(file_path)

        header_match = _PROC_HEADER_PATTERN.search(content)
        if not header_match:
            logger.warning("无法识别存储过程名: %s", file_path.name)
            return ProcedureIndicatorInfo()

        full_name = header_match.group(1).strip().upper()
        proc_name = full_name.split(".")[-1] if "." in full_name else full_name

        step_order = _PROC_STEP_MAP.get(proc_name, 0)

        index_type = self._determine_index_type(proc_name)
        config_table = self._extract_config_table(content)
        target_table = self._extract_target_table(content)
        source_tables = self._extract_source_tables(content)

        return ProcedureIndicatorInfo(
            proc_name=proc_name,
            step_order=step_order,
            index_type=index_type,
            config_table=config_table,
            target_table=target_table,
            source_tables=source_tables,
        )

    def _determine_index_type(self, proc_name: str) -> str:
        if "GL" in proc_name.upper():
            return "1"
        if "BASE" in proc_name.upper():
            return "1"
        if "DERIVE" in proc_name.upper():
            return "2"
        if "BRCHSUM" in proc_name.upper():
            return ""
        return ""

    def _extract_config_table(self, content: str) -> str:
        match = _CONFIG_TABLE_PATTERN.search(content)
        if match:
            return match.group(1).upper()
        return ""

    def _extract_target_table(self, content: str) -> str:
        targets: list[str] = []
        for match in _INSERT_INTO_PATTERN.finditer(content):
            raw_name = match.group(1).strip()
            name = raw_name.split(".")[-1] if "." in raw_name else raw_name
            name_upper = name.upper()
            if _is_skip_table(name_upper):
                continue
            if name_upper not in targets:
                targets.append(name_upper)
        if targets:
            return targets[0]
        if "FDL_IDX_TMP_INDEX_DATA_DTLBRCH" in content.upper():
            return "FDL_IDX_TMP_INDEX_DATA_DTLBRCH"
        if "FDL_IDX_TMP_INDEX_DATA_ALLBRCH" in content.upper():
            return "FDL_IDX_TMP_INDEX_DATA_ALLBRCH"
        return ""

    def _extract_source_tables(self, content: str) -> list[str]:
        tables: list[str] = []
        for match in _FROM_JOIN_PATTERN.finditer(content):
            raw_name = match.group(1).strip()
            name = raw_name.split(".")[-1] if "." in raw_name else raw_name
            name_upper = name.upper()
            if _is_skip_table(name_upper):
                continue
            if name_upper not in tables:
                tables.append(name_upper)
        return tables

    @staticmethod
    def _build_col_index(
        header: list[str], expected_columns: tuple[str, ...]
    ) -> dict[str, int]:
        col_index: dict[str, int] = {}
        for i, col_name in enumerate(header):
            cleaned = col_name.strip()
            if cleaned in expected_columns:
                col_index[cleaned] = i
        for idx, name in enumerate(expected_columns):
            if name not in col_index:
                col_index[name] = idx
        return col_index
